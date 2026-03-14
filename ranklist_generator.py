import pandas as pd
import re
import numpy as np
from typing import Dict, List, Optional
import argparse
import glob
import os

# --- CONFIGURATION: DEFINE THE MAPPING ---
# This dictionary maps the REQUIRED TEMPLATE HEADINGS (Key) to the
# INTERNAL COLUMN NAMES from the PROCESSED RANKLIST (Value).
# NOTE: We skip the formula columns B ('Competitive Rank') and C ('Normal Rank')
# NOTE: Score column will be detected dynamically based on the exam
FINAL_TEMPLATE_MAPPING_BASE: Dict[str, str] = {
    # TEMPLATE HEADING        : INTERNAL COLUMN NAME
    'Reg. No': 'Username',                    # Student registration number (Username from raw data)
    'Batch': 'Batch ID',                      # Using extracted batch ID (e.g., 7C1) without year/suffix
    'Name': 'Full Name',                      # The concatenated name
    'Correct Answers': 'No. Of Correct Answers',  # Exactly as in raw data
    'Incorrect Answers': 'No. Of incorrect Answers',  # Exactly as in raw data
    'Unanswered Questions': 'No. Of unanswered Questions',  # Exactly as in raw data
    'Score (60)': 'DYNAMIC_SCORE_COLUMN',    # Will be detected dynamically
    'Percentage (%)': 'Percentage'           # Using raw percentage
}

# Define which rows to skip when reading the template and the raw file
TEMPLATE_SKIP_ROWS = 2 # Skip the header rows (Title and Column Names) in your template
DEFAULT_RAW_GLOB = "*scores_report*.csv"
RAW_DATA_FILE: Optional[str] = None
# PERMANENT TEMPLATE FILE - This is your master template that will be used automatically
PERMANENT_TEMPLATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "manager_template (5).xlsx")


# --- DATA PROCESSING FUNCTIONS (Same as before, but combined) ---

def detect_score_column(df: pd.DataFrame) -> str:
    """
    Return 'Raw Score' as the score column to use.
    Raw Score is the total score and should be used for all exams.
    """
    if 'Raw Score' not in df.columns:
        raise ValueError("'Raw Score' column not found in the data. This column is required.")
    
    return 'Raw Score'


def process_raw_ranklist(file_path: str) -> pd.DataFrame:
    """Performs data cleaning, concatenation, extraction, and sorting by performance."""
    try:
        # Read CSV with error handling for malformed lines
        # on_bad_lines='skip' will skip problematic lines instead of crashing
        # warn will notify us which lines were skipped
        df = pd.read_csv(file_path, on_bad_lines='warn', engine='python')
    except FileNotFoundError:
        raise FileNotFoundError(f"Error: Raw ranklist file not found at {file_path}")

    # 1. CONCATENATE FIRST NAME (B) AND LAST NAME (C)
    df['First Name'] = df['First Name'].fillna('')
    df['Last Name'] = df['Last Name'].fillna('')
    df['Full Name'] = df['First Name'].astype(str) + ' ' + df['Last Name'].astype(str)

    # 2. EXTRACT BATCH IDENTIFIER
    def extract_batch_id(batch_str):
        if pd.isna(batch_str):
            return np.nan
        batch_str = str(batch_str).strip()
        
        # Handle comma-separated values by looking for patterns like '7C3', '7G2', etc.
        # Split by comma and look for batch ID patterns in each part
        import re
        parts = [part.strip() for part in batch_str.split(',')]
        
        for part in parts:
            # Look for the most common batch format (e.g., 7C3, 7C4, etc.)
            # Pattern: digits followed by letters followed by digits (e.g., 7C3, 7G2, etc.)
            batch_pattern = r'[0-9]+[A-Za-z]+[0-9]+'
            batch_match = re.search(batch_pattern, part)
            if batch_match:
                return batch_match.group()
        
        # If no pattern matches, try to extract the first alphanumeric sequence
        match = re.match(r'([A-Za-z0-9]+)', batch_str)
        return match.group(1) if match else np.nan

    df['Batch ID'] = df['Batch'].apply(extract_batch_id)

    # 3. CONVERT TIME TAKEN TO SECONDS FOR SORTING
    def time_to_seconds(time_str):
        """Convert time format 'H:MM:SS' or 'M:SS' to total seconds"""
        if pd.isna(time_str):
            return float('inf')  # Put missing times at the end
        try:
            parts = str(time_str).split(':')
            if len(parts) == 3:  # H:MM:SS format
                hours, minutes, seconds = map(int, parts)
                return hours * 3600 + minutes * 60 + seconds
            elif len(parts) == 2:  # MM:SS format
                minutes, seconds = map(int, parts)
                return minutes * 60 + seconds
            else:
                return float('inf')
        except:
            return float('inf')
    
    df['Time in Seconds'] = df['Time Taken'].apply(time_to_seconds)

    # 4. PERFORMANCE-BASED SORTING
    # Sort by: Raw Score (desc), Incorrect Answers (asc), Time Taken (asc)
    df = df.sort_values(
        by=['Raw Score', 'No. Of incorrect Answers', 'Time in Seconds'],
        ascending=[False, True, True],  # High score first, low incorrect first, fast time first
        na_position='last'
    )

    return df.reset_index(drop=True)


def get_custom_title() -> str:
    """Prompts the user to enter a custom title for the ranklist."""
    print("\n--- Custom Title for Ranklist ---")
    title_input = input(
        "Enter the title text for the ranklist (or leave blank for no title): "
    ).strip()
    return title_input


def get_output_filename() -> str:
    """Prompts the user to enter a custom filename for output files."""
    print("\n--- Output Filename ---")
    filename_input = input(
        "Enter the filename for output files (without extension): "
    ).strip()
    
    # If empty, use default
    if not filename_input:
        return "ranklist_output"
    
    # Remove any file extension if user accidentally included it
    if '.' in filename_input:
        filename_input = filename_input.rsplit('.', 1)[0]
    
    # Remove invalid filename characters
    import re
    filename_input = re.sub(r'[<>:"/\\|?*]', '_', filename_input)
    
    return filename_input


def get_batches_to_filter(df: pd.DataFrame) -> List[str]:
    """Prompts the user to select batches for filtering."""
    available_batches = df['Batch ID'].dropna().unique().tolist()
    available_batches.sort()
    
    # Simple interactive CLI for demonstration
    print("\n--- Available Batches for Filtering ---")
    print(", ".join(available_batches))
    print("---------------------------------------")

    user_input = input(
        "Enter batch IDs to filter (e.g., 7C7, 7C8) or leave blank to skip filtering: "
    ).strip()

    if not user_input:
        return []
    
    selected_batches = [b.strip().upper() for b in user_input.split(',') if b.strip()]
    valid_batches = [b for b in selected_batches if b in available_batches]
    
    if not valid_batches and selected_batches:
        print("\nNo valid batches selected. Proceeding without filtering.")
        return []

    return valid_batches


def integrate_to_template(processed_df: pd.DataFrame, template_path: str, mapping: Dict[str, str], skip_rows: int, custom_title: str = "", out_base_name: Optional[str] = None, out_dir: Optional[str] = None, make_pdf: bool = True):
    """
    Filters, selects, renames, and appends data to the permanent template file.
    """
    template_headings = list(mapping.keys())
    source_columns = list(mapping.values())
    
    # 1. Prepare the new data according to the template structure
    final_data_to_add = processed_df[source_columns].copy()
    final_data_to_add.columns = template_headings

    # 2. Read the existing template to understand structure
    try:
        # Load template workbook and sheet
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("The 'openpyxl' library is required. Please install it with 'pip install openpyxl'.")
    
    try:
        # Load template workbook and sheet
        template_wb = openpyxl.load_workbook(template_path)
        out_ws = template_wb.active
        
        # Determine output directory and base name
        out_dir = out_dir or os.getcwd()
        if not os.path.isdir(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        template_basename = os.path.splitext(os.path.basename(template_path))[0]
        base = out_base_name or f"{template_basename}_copy"

        out_xlsx_path = os.path.join(out_dir, f"{base}.xlsx")
        out_csv_path = os.path.join(out_dir, f"{base}.csv")
        out_pdf_path = os.path.join(out_dir, f"{base}.pdf")

        # 1. Clear existing data rows in the template from row skip_rows+1 onwards
        # This prevents leftover data when the new filtered DataFrame is shorter than the template's previous data
        max_row = out_ws.max_row
        if max_row > skip_rows:
            # Delete rows starting from skip_rows+1 up to max_row
            out_ws.delete_rows(skip_rows + 1, max_row - skip_rows)

        # 2. Add custom title if provided
        if custom_title:
            out_ws.merge_cells('B1:E1')
            title_cell = out_ws.cell(row=1, column=2)  # B1
            title_cell.value = custom_title
        
        # 3. Add data rows starting from row 3 (skip_rows + 1)
        start_row = skip_rows + 1
        for row_num, (idx, data_row) in enumerate(final_data_to_add.iterrows()):
            excel_row = start_row + row_num
            
            # Column A: Sequential number
            out_ws.cell(row=excel_row, column=1, value=row_num + 1)
            
            # Column B: Competitive Rank formula
            if row_num == 0:
                out_ws.cell(row=excel_row, column=2, value=1)
            else:
                formula = f"=IF(AND(C{excel_row-1}=C{excel_row},G{excel_row-1}=G{excel_row}),B{excel_row-1},A{excel_row})"
                out_ws.cell(row=excel_row, column=2, value=formula)
            
            # Column C: Normal Rank formula
            if row_num == 0:
                out_ws.cell(row=excel_row, column=3, value=1)
            else:
                formula = f"=IF(J{excel_row}=J{excel_row-1},C{excel_row-1},C{excel_row-1}+1)"
                out_ws.cell(row=excel_row, column=3, value=formula)
            
            # Columns D-K: Data from processed dataframe
            for col_idx, (header, value) in enumerate(zip(template_headings, data_row), start=4):
                cell = out_ws.cell(row=excel_row, column=col_idx)
                if pd.notna(value):
                    if header == 'Percentage (%)' and isinstance(value, (int, float, np.integer, np.floating)):
                        cell.value = f"{value}%"
                    elif isinstance(value, (int, float, np.integer, np.floating)):
                        cell.value = float(value) if isinstance(value, (float, np.floating)) else int(value)
                    else:
                        cell.value = str(value)
                else:
                    cell.value = None
            
            # Columns L-O: Empty columns
            for col_idx in range(12, 16):
                out_ws.cell(row=excel_row, column=col_idx, value=None)
        
        # Save XLSX
        template_wb.save(out_xlsx_path)
        template_wb.close()
        print(f"✅ XLSX saved at: {out_xlsx_path}")
        
        # We'll generate CSV and PDF together using calculated rank values
        # Load the workbook we just created to read data
        calc_wb = openpyxl.load_workbook(out_xlsx_path)
        calc_ws = calc_wb.active
        
        # Build data with calculated ranks (for both CSV and PDF)
        csv_data = []
        
        # Header row (row 2 from Excel) - columns B to K
        header_row = []
        for col_idx in range(2, 12):  # Columns B-K (2-11)
            cell_val = calc_ws.cell(row=2, column=col_idx).value
            # Clean header text (remove newlines for CSV)
            cell_str = str(cell_val).replace('\n', ' ').replace('\r', ' ').strip() if cell_val is not None else ''
            header_row.append(cell_str)
        csv_data.append(header_row)
        
        # Calculate ranks and build data rows
        prev_comp_rank = 1
        prev_normal_rank = 1
        
        for row_idx in range(3, calc_ws.max_row + 1):
            # Check if row has data
            reg_no = calc_ws.cell(row=row_idx, column=4).value
            batch = calc_ws.cell(row=row_idx, column=5).value
            name = calc_ws.cell(row=row_idx, column=6).value
            score = calc_ws.cell(row=row_idx, column=10).value
            
            if all(v is None for v in [reg_no, batch, name, score]):
                break
            
            data_row = []
            
            # Calculate Competitive Rank and Normal Rank
            if row_idx == 3:
                comp_rank = 1
                normal_rank = 1
            else:
                # Calculate Normal Rank
                prev_percentage = calc_ws.cell(row=row_idx - 1, column=11).value
                curr_percentage = calc_ws.cell(row=row_idx, column=11).value
                
                if prev_percentage == curr_percentage:
                    normal_rank = prev_normal_rank
                else:
                    normal_rank = prev_normal_rank + 1
                
                # Calculate Competitive Rank
                prev_correct = calc_ws.cell(row=row_idx - 1, column=7).value
                curr_correct = calc_ws.cell(row=row_idx, column=7).value
                seq_num = row_idx - 2
                
                if prev_normal_rank == normal_rank and prev_correct == curr_correct:
                    comp_rank = prev_comp_rank
                else:
                    comp_rank = seq_num
            
            prev_comp_rank = comp_rank
            prev_normal_rank = normal_rank
            
            # Add ranks to row
            data_row.append(str(int(comp_rank)))
            data_row.append(str(int(normal_rank)))
            
            # Add other columns (D-K: columns 4-11)
            for col_idx in range(4, 12):
                cell_val = calc_ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    if isinstance(cell_val, (int, float, np.integer, np.floating)):
                        # Column 11 is Percentage - add % symbol
                        if col_idx == 11:
                            if isinstance(cell_val, (float, np.floating)) and cell_val % 1 == 0:
                                data_row.append(f"{int(cell_val)}%")
                            else:
                                data_row.append(f"{cell_val}%")
                        elif isinstance(cell_val, (float, np.floating)) and cell_val % 1 == 0:
                            data_row.append(str(int(cell_val)))
                        else:
                            data_row.append(str(cell_val))
                    else:
                        data_row.append(str(cell_val))
                else:
                    data_row.append('')
            
            csv_data.append(data_row)
        
        # Save CSV with calculated ranks
        import csv
        with open(out_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(csv_data)
        print(f"✅ CSV saved at: {out_csv_path}")

        # Create PDF with template colors
        if make_pdf:
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph
                from reportlab.lib.units import inch
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_LEFT
                
                # A4 dimensions in points: 595.28 x 841.89
                # Using point-based measurements for precise control
            except ImportError:
                print("Could not create PDF: The 'reportlab' library is required.")
                print(f"PDF was not created. Intended path: {out_pdf_path}")
                calc_wb.close()
                return

            try:
                # Register Calibri font (fallback to Helvetica if Calibri not available)
                calibri_available = False
                try:
                    # Try to register Calibri font - should be available on Windows
                    # Common Calibri font paths on Windows
                    calibri_paths = [
                        r'C:\Windows\Fonts\calibri.ttf',
                        r'C:\Windows\Fonts\Calibri.ttf',
                    ]
                    for font_path in calibri_paths:
                        if os.path.exists(font_path):
                            pdfmetrics.registerFont(TTFont('Calibri', font_path))
                            calibri_available = True
                            break
                    
                    if calibri_available:
                        # Try to register Calibri Bold
                        calibri_bold_paths = [
                            r'C:\Windows\Fonts\calibrib.ttf',
                            r'C:\Windows\Fonts\CalibriB.ttf',
                        ]
                        for font_path in calibri_bold_paths:
                            if os.path.exists(font_path):
                                pdfmetrics.registerFont(TTFont('Calibri-Bold', font_path))
                                break
                        else:
                            # If Calibri Bold not found, use regular Calibri for bold too
                            pdfmetrics.registerFont(TTFont('Calibri-Bold', calibri_paths[0]))
                except Exception:
                    calibri_available = False
                
                # Set font names based on availability
                if calibri_available:
                    font_regular = 'Calibri'
                    font_bold = 'Calibri-Bold'
                    print("Using Calibri font for PDF")
                else:
                    font_regular = 'Helvetica'
                    font_bold = 'Helvetica-Bold'
                    print("Calibri font not found. Using Helvetica as fallback.")
                # Build PDF data from csv_data (reuse calculated ranks)
                pdf_data = []
                
                # Title row (if custom_title is provided)
                if custom_title:
                    title_row = [custom_title] + [''] * 9
                    pdf_data.append(title_row)
                
                # Add header row with line breaks for Competitive Rank and Normal Rank
                pdf_header_row = [
                    'Competitive\nRank',      # Line break for Competitive Rank
                    'Normal\nRank',           # Line break for Normal Rank
                    csv_data[0][2],           # Reg. No
                    csv_data[0][3],           # Batch
                    csv_data[0][4],           # Name
                    'Correct\nAnswers',       # Line break for Correct Answers
                    'Incorrect\nAnswers',     # Line break for Incorrect Answers
                    'Unanswered\nQuestions',  # Line break for Unanswered Questions
                    'Score',                  # Score without (60)
                    'Percentage\n(%)',        # Line break for Percentage (%)
                ]
                pdf_data.append(pdf_header_row)
                
                # Add data rows (skip header from csv_data)
                pdf_data.extend(csv_data[1:])
                
                # Create ParagraphStyle for Name column to ensure proper wrapping
                name_style = ParagraphStyle(
                    'NameStyle',
                    fontName=font_regular,
                    fontSize=9,
                    alignment=TA_LEFT,
                    leading=11,  # Line spacing
                    wordWrap='LTR',  # Left-to-right word wrap
                    splitLongWords=True  # Force breaking of long words
                )
                
                # Wrap Name column (index 4) in Paragraph objects for proper text breaking
                # Start from row 2 if custom_title exists (skip title and header), else row 1 (skip header)
                start_row = 2 if custom_title else 1
                for i in range(start_row, len(pdf_data)):
                    name_text = str(pdf_data[i][4]) if pdf_data[i][4] else ''  # Handle None values
                    pdf_data[i][4] = Paragraph(name_text, name_style)
                
                calc_wb.close()
                
                # Create PDF with portrait orientation using exact point measurements
                # A4: 595.28pt x 841.89pt, Margins: 35pt L/R, 55pt T, 45pt B
                doc = SimpleDocTemplate(out_pdf_path, pagesize=A4,
                                      leftMargin=35, rightMargin=35,
                                      topMargin=55, bottomMargin=45)
                
                # Define column widths in points (usable width: 525pt)
                # Increased Name column width to prevent text overflow with wrapping
                col_widths = [
                    55,   # Competitive Rank (reduced slightly)
                    43,   # Normal Rank (reduced slightly)
                    52,   # Reg. No (reduced slightly)
                    30,   # Batch (reduced slightly)
                    125,  # Name (increased from 105pt to prevent overflow, left-aligned)
                    43,   # Correct Answers (reduced slightly)
                    51,   # Incorrect Answers (reduced slightly)
                    60,   # Unanswered Questions (reduced slightly)
                    35,   # Score (reduced slightly)
                    55,   # Percentage (%) (reduced slightly)
                ]  # Total: 525pt (fits perfectly in usable width)
                
                # Create table with specified column widths
                table = Table(pdf_data, colWidths=col_widths)
                
                # Define colors from specification
                title_yellow = colors.HexColor('#FAC516')  # Yellow background for title
                header_red = colors.HexColor('#EB4D5D')    # Red background for column headers
                rank_gray = colors.HexColor('#D6DCE4')     # Gray background for rank columns
                
                # Build table style based on whether we have a title row
                if custom_title:
                    # With title row: row 0 is title, row 1 is headers, row 2+ is data
                    style = TableStyle([
                        # Grid
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                        # Title row styling (row 0) - yellow background, merged across all columns
                        ('BACKGROUND', (0, 0), (-1, 0), title_yellow),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('FONTNAME', (0, 0), (-1, 0), font_bold),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),  # Title font size
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                        ('SPAN', (0, 0), (-1, 0)),  # Merge all columns in row 0
                        # Column header row styling (row 1) - red background with black text
                        ('BACKGROUND', (0, 1), (-1, 1), header_red),
                        ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),  # Black text in red header
                        ('FONTNAME', (0, 1), (-1, 1), font_bold),
                        ('FONTSIZE', (0, 1), (-1, 1), 9),  # Header font size
                        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
                        ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
                        # Competitive Rank and Normal Rank columns (0-1) - gray background for data rows
                        ('BACKGROUND', (0, 2), (1, -1), rank_gray),
                        # Data rows styling (row 2+)
                        ('FONTNAME', (0, 2), (-1, -1), font_regular),
                        ('FONTSIZE', (0, 2), (-1, -1), 9),  # Data font size
                        ('ALIGN', (0, 2), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 2), (-1, -1), 'MIDDLE'),
                        # Name column (column 4, index 4) - LEFT alignment for data rows only
                        ('ALIGN', (4, 2), (4, -1), 'LEFT'),
                        # Word wrap for all cells
                        ('WORDWRAP', (0, 0), (-1, -1), True),
                        # Reduced padding for smaller, uniform cells
                        ('LEFTPADDING', (0, 0), (-1, -1), 3),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        # Special padding for header row
                        ('TOPPADDING', (0, 1), (-1, 1), 4),
                        ('BOTTOMPADDING', (0, 1), (-1, 1), 4),
                    ])
                    # Set repeatRows to repeat title and header rows (rows 0-1) on every page
                    table.repeatRows = 2
                else:
                    # Without title row: row 0 is headers, row 1+ is data
                    style = TableStyle([
                        # Grid
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                        # Header row styling (row 0) - red background with black text
                        ('BACKGROUND', (0, 0), (-1, 0), header_red),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Black text in red header
                        ('FONTNAME', (0, 0), (-1, 0), font_bold),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),  # Header font size
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                        # Competitive Rank and Normal Rank columns (0-1) - gray background for data rows
                        ('BACKGROUND', (0, 1), (1, -1), rank_gray),
                        # Data rows styling (row 1+)
                        ('FONTNAME', (0, 1), (-1, -1), font_regular),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),  # Data font size
                        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                        # Name column (column 4, index 4) - LEFT alignment for data rows only
                        ('ALIGN', (4, 1), (4, -1), 'LEFT'),
                        # Word wrap for all cells
                        ('WORDWRAP', (0, 0), (-1, -1), True),
                        # Reduced padding for smaller, uniform cells
                        ('LEFTPADDING', (0, 0), (-1, -1), 3),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                        # Special padding for header row
                        ('TOPPADDING', (0, 0), (-1, 0), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
                    ])
                    # Set repeatRows to repeat header row (row 0) on every page
                    table.repeatRows = 1
                
                table.setStyle(style)
                elements = [table]
                doc.build(elements)
                print(f"✅ PDF saved at: {out_pdf_path}")
            except Exception as e:
                print(f"Failed to write PDF at {out_pdf_path}: {e}")
                import traceback
                traceback.print_exc()
                return
        
    except FileNotFoundError:
        print(f"\nTemplate file not found at {template_path}. Please ensure it exists.")
        return
    except Exception as e:
        print(f"\nAn error occurred during template integration: {e}")
        import traceback
        traceback.print_exc()
        return

    print(f"\n✅ Success! New filtered data has been processed.")
    print(f"Number of new records added: {len(final_data_to_add)}")
    
    # Check if the target student is in the final data
    if 'Username' in final_data_to_add.columns:
        target_students = final_data_to_add[
            (final_data_to_add['Username'] == 'F25070871') | 
            (final_data_to_add['Full Name'].str.contains('zayan Muhammed', case=False, na=False))
        ]
        if len(target_students) > 0:
            print(f"✅ Target student 'F25070871 zayan Muhammed' is included in the final output.")
        else:
            print(f"❌ Target student 'F25070871 zayan Muhammed' is NOT in the final output.")
            print(f"   Available usernames in final data: {final_data_to_add['Username'].head(5).tolist() if 'Username' in final_data_to_add.columns else 'Username column not found'}")
            print(f"   Available names in final data: {final_data_to_add['Full Name'].head(5).tolist() if 'Full Name' in final_data_to_add.columns else 'Full Name column not found'}")


def main():
    """Main function to run the entire ranklist generation process."""
    try:
        # Determine raw data file: CLI arg > auto-detect > error
        parser = argparse.ArgumentParser(description="Generate ranklist into template.")
        parser.add_argument("raw_file", nargs="?", help="Path to the raw scores CSV file")
        parser.add_argument("--template", "-t", dest="template", help="Path to the template (.xlsx or .csv)")
        parser.add_argument("--out-name", dest="out_name", help="Base name for output CSV and PDF (without extension)")
        parser.add_argument("--out-dir", dest="out_dir", help="Directory to save output files")
        parser.add_argument("--no-pdf", dest="no_pdf", action='store_true', help="Do not generate PDF output")
        args = parser.parse_args()

        raw_file_to_use = None
        if args.raw_file:
            raw_file_to_use = args.raw_file
        else:
            # Search workspace for candidate files
            candidates = glob.glob(DEFAULT_RAW_GLOB)
            candidates += glob.glob(os.path.join(os.getcwd(), DEFAULT_RAW_GLOB))
            
            # Deduplicate by normalizing paths to absolute paths
            unique_candidates = {}
            for f in candidates:
                abs_path = os.path.abspath(f)
                if abs_path not in unique_candidates:
                    unique_candidates[abs_path] = f
            candidates = list(unique_candidates.keys())
            
            # Sort by modification time (newest first) and keep only last 5
            if candidates:
                candidates_with_time = [(f, os.path.getmtime(f)) for f in candidates]
                candidates_with_time.sort(key=lambda x: x[1], reverse=True)  # Sort by time, newest first
                candidates = [f for f, _ in candidates_with_time[:5]]  # Keep only last 5 newest files
            
            if len(candidates) == 1:
                raw_file_to_use = candidates[0]
            elif len(candidates) > 1:
                # Interactive selection - showing only last 5 newest files
                print("\n📄 Last 5 recently added files:")
                for i, c in enumerate(candidates, start=1):
                    # Get file modification time for display
                    mod_time = os.path.getmtime(c)
                    from datetime import datetime
                    time_str = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')
                    print(f"{i}. {os.path.basename(c)} (Modified: {time_str})")
                choice = input("\nEnter the number of the file to use (or press Enter to cancel): ").strip()
                if choice.isdigit() and 1 <= int(choice) <= len(candidates):
                    raw_file_to_use = candidates[int(choice) - 1]
            else:
                raw_file_to_use = None

        if not raw_file_to_use:
            raise FileNotFoundError(
                "Error: Raw ranklist file not specified and no candidate '*scores_report*.csv' files were found in the current directory."
            )

        print(f"1. Processing raw ranklist from: {raw_file_to_use}")
        processed_df = process_raw_ranklist(raw_file_to_use)
        
        # Check if target student exists in the original processed data
        target_student = processed_df[(processed_df['Username'] == 'F25070871') | 
                                    (processed_df['Full Name'].str.contains('zayan Muhammed', case=False, na=False))]
        if len(target_student) > 0:
            print(f"✅ Target student 'F25070871 zayan Muhammed' found in raw data ({len(target_student)} record(s)).")
        else:
            print(f"❌ Target student 'F25070871 zayan Muhammed' NOT found in raw data.")
            print(f"   Total records in raw data: {len(processed_df)}")
            print(f"   Available usernames in raw data: {processed_df['Username'].head(10).tolist()}")
            print(f"   Available names in raw data: {processed_df['Full Name'].head(10).tolist()}")

        # Detect the score column dynamically
        score_column = detect_score_column(processed_df)
        print(f"2. Detected score column: '{score_column}'")
        
        # Create the final mapping with the detected score column
        FINAL_TEMPLATE_MAPPING = FINAL_TEMPLATE_MAPPING_BASE.copy()
        FINAL_TEMPLATE_MAPPING['Score (60)'] = score_column

        # STEP 3: Get custom title
        custom_title = get_custom_title()

        # STEP 4: Get output filename
        if not args.out_name:
            output_filename = get_output_filename()
        else:
            output_filename = args.out_name

        # STEP 5: Interactive Filtering
        batches_to_keep = get_batches_to_filter(processed_df)

        if batches_to_keep:
            # Before filtering, check if any important records might be lost
            original_count = len(processed_df)
            processed_df = processed_df[processed_df['Batch ID'].isin(batches_to_keep)].copy()
            filtered_count = len(processed_df)
            print(f"Filtering complete. {filtered_count} records remaining (from {original_count}).")
            
            # Check if the specific student is in the filtered data
            target_student = processed_df[(processed_df['Username'] == 'F25070871') | 
                                        (processed_df['Full Name'].str.contains('zayan Muhammed', case=False, na=False))]
            if len(target_student) == 0:
                print("⚠️  Warning: The student 'F25070871 zayan Muhammed' was filtered out by batch filtering.")
                print("   Available batches in data:", processed_df['Batch ID'].unique())
                print("   Batches that were selected for filtering:", batches_to_keep)
        else:
            print("Filtering skipped. All records will be considered for the template.")
            
            # Check if the specific student exists in the data
            target_student = processed_df[(processed_df['Username'] == 'F25070871') | 
                                        (processed_df['Full Name'].str.contains('zayan Muhammed', case=False, na=False))]
            if len(target_student) > 0:
                print(f"✅ Found student 'F25070871 zayan Muhammed' in the dataset ({len(target_student)} record(s)).")
            else:
                print("⚠️  Warning: Could not find student 'F25070871 zayan Muhammed' in the dataset.")
                print("   Available usernames in data:", processed_df['Username'].head(10).tolist())
                print("   Available names in data:", processed_df['Full Name'].head(10).tolist())

        # STEP 6: Use the current directory as output directory
        output_dir = os.getcwd()
        
        # STEP 7: Integrate into the Permanent Template (without modifying the original)
        template_to_use = args.template if args.template else PERMANENT_TEMPLATE_FILE
        print(f"\n6. Integrating data into template (original left unchanged). Template source: {template_to_use}")
        integrate_to_template(
            processed_df=processed_df,
            template_path=template_to_use,
            mapping=FINAL_TEMPLATE_MAPPING,
            skip_rows=TEMPLATE_SKIP_ROWS,
            custom_title=custom_title,
            out_base_name=output_filename,
            out_dir=output_dir,
            make_pdf=(not args.no_pdf)
        )

    except Exception as e:
        print(f"\n🛑 An error occurred during execution: {e}")


if __name__ == "__main__":
    main()