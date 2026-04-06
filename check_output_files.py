import openpyxl
import csv

# Check the generated output files
print("=" * 80)
print("CHECKING GENERATED OUTPUT FOR MISSING ROLL NUMBERS")
print("=" * 80)

target_numbers = ['F26102027', 'F26101393', 'F26100131', 'F26100902', 'F25091944', 'F25092126', 'F24081853', 'F26101469']

# Check CSV file first (easier to read)
csv_file = "10.csv"
print(f"\n1. Checking {csv_file}:")
try:
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        
    print(f"   Total rows (including header): {len(rows)}")
    print(f"   Data rows: {len(rows) - 1}")
    
    # Check for each target number
    for num in target_numbers:
        found = False
        for i, row in enumerate(rows, 1):
            # Username is in column C (index 2) based on the template structure
            # But after ranks are added, it might be in different position
            for cell in row:
                if num in cell:
                    print(f"   ✅ {num} found in row {i}: {row}")
                    found = True
                    break
            if found:
                break
        
        if not found:
            print(f"   ❌ {num} NOT FOUND in CSV")

except Exception as e:
    print(f"   Error reading CSV: {e}")

# Check XLSX file
xlsx_file = "10.xlsx"
print(f"\n2. Checking {xlsx_file}:")
try:
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    
    print(f"   Total rows: {ws.max_row}")
    
    # Check for each target number
    for num in target_numbers:
        found = False
        for row_num in range(3, ws.max_row + 1):  # Start from row 3 (data starts after headers)
            # Reg. No is typically in column D (column 4)
            cell_value = ws.cell(row=row_num, column=4).value
            if cell_value and num in str(cell_value):
                print(f"   ✅ {num} found at row {row_num}")
                # Print the full row data
                row_data = []
                for col in range(4, 12):
                    val = ws.cell(row=row_num, column=col).value
                    row_data.append(str(val) if val else '')
                print(f"      Data: {row_data}")
                found = True
                break
        
        if not found:
            print(f"   ❌ {num} NOT FOUND in XLSX")
    
    wb.close()

except Exception as e:
    print(f"   Error reading XLSX: {e}")

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
