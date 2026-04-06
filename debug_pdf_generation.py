import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table
import csv

print("=" * 80)
print("DEBUGGING PDF GENERATION")
print("=" * 80)

# Load the XLSX file
xlsx_file = "10.xlsx"
print(f"\n1. Loading {xlsx_file}...")
wb = openpyxl.load_workbook(xlsx_file)
ws = wb.active
print(f"   Total rows in XLSX: {ws.max_row}")

# Build CSV data the same way ranklist_generator.py does it
csv_data = []

# Header row (row 2 from Excel) - columns B to K
header_row = []
for col_idx in range(2, 12):  # Columns B-K (2-11)
    cell_val = ws.cell(row=2, column=col_idx).value
    cell_str = str(cell_val).replace('\n', ' ').replace('\r', ' ').strip() if cell_val is not None else ''
    header_row.append(cell_str)
csv_data.append(header_row)

print(f"\n2. Header row: {header_row}")

# Build data rows
print(f"\n3. Reading data rows from XLSX...")
data_row_count = 0
for row_idx in range(3, ws.max_row + 1):
    # Check if row has data
    reg_no = ws.cell(row=row_idx, column=4).value
    batch = ws.cell(row=row_idx, column=5).value
    name = ws.cell(row=row_idx, column=6).value
    score = ws.cell(row=row_idx, column=10).value
    
    if all(v is None for v in [reg_no, batch, name, score]):
        print(f"   Stopping at row {row_idx} - no more data")
        break
    
    data_row = []
    data_row.append(str(row_idx - 2))  # Sequential number as rank for now
    data_row.append(str(row_idx - 2))
    
    # Add other columns (D-K: columns 4-11)
    for col_idx in range(4, 12):
        cell_val = ws.cell(row=row_idx, column=col_idx).value
        if cell_val is not None:
            data_row.append(str(cell_val))
        else:
            data_row.append('')
    
    csv_data.append(data_row)
    data_row_count += 1

print(f"   Total data rows collected: {data_row_count}")

# Check for target numbers
target_numbers = ['F26102027', 'F26101393', 'F26100131', 'F26100902', 'F25091944', 'F25092126', 'F24081853', 'F26101469']

print(f"\n4. Checking target numbers in csv_data:")
for num in target_numbers:
    found = False
    for i, row in enumerate(csv_data):
        for cell in row:
            if num in str(cell):
                print(f"   ✅ {num} found in csv_data row {i}")
                found = True
                break
        if found:
            break
    if not found:
        print(f"   ❌ {num} NOT in csv_data")

# Now check the PDF file
print(f"\n5. Checking PDF file...")
pdf_file = "10.pdf"
try:
    # Try to extract text from PDF
    from pdfplumber import open as pdf_open
    
    with pdf_open(pdf_file) as pdf:
        print(f"   Total pages: {len(pdf.pages)}")
        
        # Extract all text from PDF
        all_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_text += text
        
        # Check for target numbers
        print(f"\n6. Checking target numbers in PDF text:")
        for num in target_numbers:
            if num in all_text:
                print(f"   ✅ {num} FOUND in PDF")
            else:
                print(f"   ❌ {num} NOT FOUND in PDF")
                
except ImportError:
    print("   pdfplumber not installed, trying alternative method...")
    try:
        import PyPDF2
        
        with open(pdf_file, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            print(f"   Total pages: {len(reader.pages)}")
            
            # Extract all text
            all_text = ""
            for page in reader.pages:
                all_text += page.extract_text()
            
            # Check for target numbers
            print(f"\n6. Checking target numbers in PDF text:")
            for num in target_numbers:
                if num in all_text:
                    print(f"   ✅ {num} FOUND in PDF")
                else:
                    print(f"   ❌ {num} NOT FOUND in PDF")
    except ImportError:
        print("   No PDF reading library available")
        print("   Install with: pip install pdfplumber OR pip install PyPDF2")
        
        # Just check if PDF file exists and its size
        import os
        if os.path.exists(pdf_file):
            size = os.path.getsize(pdf_file)
            print(f"\n   PDF file exists: {pdf_file}")
            print(f"   File size: {size / 1024:.2f} KB")
            print(f"   If file size is reasonable (>100KB), PDF likely has data")

wb.close()

print("\n" + "=" * 80)
print("DIAGNOSIS")
print("=" * 80)
print("\nIf numbers are in csv_data but NOT in PDF, the issue is in PDF rendering.")
print("If numbers are NOT in csv_data, the issue is in XLSX reading.")
